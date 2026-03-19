import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

# --- 데이터 정리 및 엑셀 생성 함수 ---
def process_data(file_guri, file_nyj):
    def clean_df(file_obj, academy_name):
        if file_obj is None: return pd.DataFrame()
        if file_obj.name.endswith('.csv'): df = pd.read_csv(file_obj)
        elif file_obj.name.endswith('.xlsx'): df = pd.read_excel(file_obj, engine='openpyxl')
        else: return pd.DataFrame()
            
        cleaned_data = []
        for index, row in df.iterrows():
            course_type = row.get('현재 진행중인 과정(근로자 혹은 실업자)을 선택해 주세요.(*)', '')
            if pd.isna(course_type) or str(course_type).strip() == '': continue
                
            if '근로자' in str(course_type):
                data = {
                    '응답일시': row.get('응답일시', ''),
                    '학원명': academy_name, '과정구분': '근로자 과정', '이름': row.get('이름을 입력해주세요.(*)', ''),
                    '1.전반적만족도': row.get('[전반적 만족도] 1. 이 훈련과정에 대해 전반적으로 만족한다.(*)', ''),
                    '2.훈련내용(실무/취업)': row.get('[훈련내용] 2. 훈련과정은 기업현장의 실무와 연계되었다.(*)', ''),
                    '3.내용일치': row.get('[내용일치] 3. HRD-Net 사이트에 제시된 수강정보(훈련목표, 내용, 방법 등)에 따라 훈련이 운영되었다.(*)', ''),
                    '4.학습방식': row.get('[학습방식] 4. 훈련과정 목적에 맞게 이론과 실습(실기)이 연계·운영되었다.(*)', ''),
                    '5.훈련시간': row.get('[훈련시간] 5. 훈련방식(이론, 실습 등)간의 시간배분이 적절하였다.(*)', ''),
                    '6.학습자료': row.get('[학습자료] 6. 훈련에 활용된 학습자료(교재, 동영상, 보조자료 등)가 적절하였다.(*)', ''),
                    '7.학습수준': row.get('[학습수준] 7. 나의 수준을 고려한 맞춤식 수업이 진행되었다.(*)', ''),
                    '8.교사/강사': row.get('[교사·강사] 8. 훈련에 대한 열의와 전문지식을 가지고 있었다.(*)', ''),
                    '9.학습평가': row.get('[학습평가] 9. 평가방법(시험, 과제 등)이 적절하였다.(*)', ''),
                    '10.피드백': row.get('[피드백] 10. 평가결과를 알려주고 부족한 부분을 보완해 주었다.(*)', ''),
                    '11.학습환경': row.get('[학습환경] 11. 학습시설(강의·실습 공간, 부대시설 등)이 적절하였다.(*)', ''),
                    '12.장비/도구': row.get('[장비 등] 12. 훈련에 필요한 장비, 도구, 재료 등이 적절하였다.(*)', ''),
                    '13.지원(경력/취업)': row.get('[경력지원] 13. 자기개발을 위해 제공된 정보(학습활동, 자격증 취득 등)가 적절하였다.(*)', ''),
                    '14.목표달성': row.get('[목표달성] 14. 나는 이 훈련과정의 목표를 달성하였다.(*)', ''),
                    '15.능력향상': row.get('[능력향상] 15. 나는 이 훈련과정을 통해 해당 분야의 직무수행능력이 향상되었다.(*)', ''),
                    '16.취업가능성(실업자)': '-',
                    '17.수강가치': row.get('[수강가치] 16. 이 훈련과정은 이 정도의 시간과 비용을 투자하여 수강할 가치가 있다.(*)', ''),
                    '18.추천여부': row.get('[추천여부] 17. 이 훈련과정을 다른 사람에게 추천하고 싶다.(*)', ''),
                    '개선요청사항': row.get('개선요청사항 (선택사항)', ''), '수강후기': row.get('수강후기 (선택사항)', '')
                }
            elif '실업자' in str(course_type): 
                data = {
                    '응답일시': row.get('응답일시', ''),
                    '학원명': academy_name, '과정구분': '실업자 과정', '이름': row.get('이름을 입력해주세요.(*).1', ''),
                    '1.전반적만족도': row.get('[전반적 만족도] 1. 이 훈련과정에 대해 전반적으로 만족한다.(*).1', ''),
                    '2.훈련내용(실무/취업)': row.get('[훈련내용] 2. 훈련과정은 취업(창업)에 필요한 실무 지식·기술로 구성되었다.(*)', ''),
                    '3.내용일치': row.get('[내용일치] 3. HRD-Net 사이트에 제시된 수강정보(훈련목표, 내용, 방법 등)에 따라 훈련이 운영되었다.(*).1', ''),
                    '4.학습방식': row.get('[학습방식] 4. 훈련과정 목적에 맞게 이론과 실습(실기)이 연계·운영되었다.(*).1', ''),
                    '5.훈련시간': row.get('[훈련시간] 5. 훈련방식(이론, 실습 등)간의 시간배분이 적절하였다.(*).1', ''),
                    '6.학습자료': row.get('[학습자료] 6. 훈련에 활용된 학습자료(교재, 동영상, 보조자료 등)가 적절하였다.(*).1', ''),
                    '7.학습수준': row.get('[학습수준] 7. 나의 수준을 고려한 맞춤식 수업이 진행되었다.(*).1', ''),
                    '8.교사/강사': row.get('[교사·강사] 8. 훈련에 대한 열의와 전문지식을 가지고 있었다.(*).1', ''),
                    '9.학습평가': row.get('[학습평가] 9. 평가방법(시험, 과제 등)이 적절하였다.(*).1', ''),
                    '10.피드백': row.get('[피드백] 10. 평가결과를 알려주고 부족한 부분을 보완해 주었다.(*).1', ''),
                    '11.학습환경': row.get('[학습환경] 11. 학습시설(강의·실습 공간, 부대시설 등)이 적절하였다.(*).1', ''),
                    '12.장비/도구': row.get('[장비 등] 12. 훈련에 필요한 장비, 도구, 재료 등이 적절하였다.(*).1', ''),
                    '13.지원(경력/취업)': row.get('[취업지원] 13. 관련 분야 취업(창업)을 위한 상담과 정보 등이 적절하였다.(*)', ''),
                    '14.목표달성': row.get('[목표달성] 14. 나는 이 훈련과정의 목표를 달성하였다.(*).1', ''),
                    '15.능력향상': row.get('[능력향상] 15. 나는 이 훈련과정을 통해 해당 분야의 직무를 수행할 수 있는 능력과 자신감이 생겼다.(*)', ''),
                    '16.취업가능성(실업자)': row.get('[취업가능성] 16. 나는 이 훈련과정을 통해 해당 분야에 취업(창업)할 가능성이 높아졌다.(*)', ''),
                    '17.수강가치': row.get('[수강가치] 17. 이 훈련과정은 이 정도의 시간과 비용을 투자하여 수강할 가치가 있다.(*)', ''),
                    '18.추천여부': row.get('[추천여부] 18. 이 훈련과정을 다른 사람에게 추천하고 싶다.(*)', ''),
                    '개선요청사항': row.get('개선요청사항 (선택사항).1', ''), '수강후기': row.get('수강후기 (선택사항).1', '')
                }
            else: continue
                
            for key in data:
                if pd.isna(data[key]): data[key] = ''
            cleaned_data.append(data)
            
        return pd.DataFrame(cleaned_data)

    df1 = clean_df(file_guri, '구리간호학원')
    df2 = clean_df(file_nyj, '남양주간호학원')
    final_df = pd.concat([df1, df2], ignore_index=True)
    
    if not final_df.empty:
        final_df = final_df.sort_values(by=['과정구분', '이름'], ascending=[True, True]).reset_index(drop=True)
        final_df.insert(0, '순번', range(1, len(final_df) + 1))
        
    return final_df

def generate_excel(final_df):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # [1. 원본 데이터 시트 생성]
        final_df.to_excel(writer, index=False, sheet_name='전체항목_원본(인쇄용)')
        wb = writer.book
        ws_raw = writer.sheets['전체항목_원본(인쇄용)']
        
        header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        for row in ws_raw.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for row in ws_raw.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        col_improve = ws_raw.max_column - 1
        col_review = ws_raw.max_column
        for row in ws_raw.iter_rows(min_row=2, min_col=col_improve, max_col=col_review):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        ws_raw.column_dimensions['A'].width = 5   
        ws_raw.column_dimensions['B'].width = 13  
        ws_raw.column_dimensions['C'].width = 11  
        ws_raw.column_dimensions['D'].width = 10  
        ws_raw.column_dimensions['E'].width = 8   
        for col_idx in range(6, ws_raw.max_column - 1):
            ws_raw.column_dimensions[get_column_letter(col_idx)].width = 11 
        ws_raw.column_dimensions[get_column_letter(col_improve)].width = 25 
        ws_raw.column_dimensions[get_column_letter(col_review)].width = 25  

        ws_raw.row_dimensions[1].height = 40 
        for r_idx, row in enumerate(ws_raw.iter_rows(min_row=2), start=2):
            improve_text = str(row[col_improve - 1].value).strip() if row[col_improve - 1].value else ""
            review_text = str(row[col_review - 1].value).strip() if row[col_review - 1].value else ""
            lines_imp = improve_text.count('\n') + (len(improve_text) // 20) + 1
            lines_rev = review_text.count('\n') + (len(review_text) // 20) + 1
            ws_raw.row_dimensions[r_idx].height = max(45, max(lines_imp, lines_rev) * 18)

        ws_raw.page_setup.orientation = ws_raw.ORIENTATION_LANDSCAPE
        ws_raw.sheet_properties.pageSetUpPr.fitToPage = True
        ws_raw.page_setup.fitToHeight = False 
        ws_raw.page_setup.fitToWidth = 1      
        ws_raw.page_margins.left = 0.2
        ws_raw.page_margins.right = 0.2
        ws_raw.page_margins.top = 0.5
        ws_raw.page_margins.bottom = 0.5

        # [2. 요약 보고서 시트 생성]
        ws_sum = wb.create_sheet(title='요약보고서(인쇄용)')
        ws_sum.append(["📊 만족도 조사 핵심 요약 보고서"])
        ws_sum['A1'].font = Font(size=16, bold=True)
        ws_sum.append([""])
        
        numeric_cols = [col for col in final_df.columns if col[0].isdigit()]
        df_numeric = final_df.copy()
        for col in numeric_cols: df_numeric[col] = pd.to_numeric(df_numeric[col], errors='coerce')
        
        overall_mean = df_numeric[numeric_cols].mean().round(2)
        mean_df = df_numeric.groupby('과정구분')[numeric_cols].mean().round(2)
        
        total_resp = len(final_df)
        overall_avg = overall_mean.mean().round(2)
        recom_cnt = final_df['18.추천여부'].astype(str).str.contains('예').sum()
        recom_rate = round((recom_cnt / total_resp) * 100, 1) if total_resp > 0 else 0
        
        ws_sum.append(["[ 💡 핵심 요약 대시보드 ]"])
        ws_sum.cell(ws_sum.max_row, 1).font = Font(bold=True)
        dashboard_text = f"  ▶ 총 응답자: {total_resp}명      |      ▶ 전체 평균 만족도: {overall_avg}점      |      ▶ 지인 추천 의향: {recom_rate}%"
        ws_sum.append([dashboard_text])
        ws_sum.merge_cells(start_row=ws_sum.max_row, start_column=1, end_row=ws_sum.max_row, end_column=5)
        ws_sum.cell(ws_sum.max_row, 1).font = Font(bold=True, size=11, color="0054FF")
        ws_sum.append([""])
        
        item_means = overall_mean.sort_values(ascending=False)
        top3 = item_means.head(3)
        bottom3 = item_means.tail(3)
        warnings = item_means[item_means < 6.0]
        
        ws_sum.append(["[ 🏆 학원 강점 (만족도 Top 3) ]"])
        ws_sum.cell(ws_sum.max_row, 1).font = Font(bold=True, color="0070C0")
        for i, (idx, val) in enumerate(top3.items(), 1):
            name = idx.split('.', 1)[1] if '.' in idx else idx
            ws_sum.append([f"  {i}위. {name} ({val}점)"])
            ws_sum.merge_cells(start_row=ws_sum.max_row, start_column=1, end_row=ws_sum.max_row, end_column=5)
        ws_sum.append([""])
        
        ws_sum.append(["[ 🚨 시급한 개선점 (만족도 Bottom 3) ]"])
        ws_sum.cell(ws_sum.max_row, 1).font = Font(bold=True, color="FF0000")
        for i, (idx, val) in enumerate(bottom3.items(), 1):
            name = idx.split('.', 1)[1] if '.' in idx else idx
            ws_sum.append([f"  {i}위. {name} ({val}점)"])
            ws_sum.merge_cells(start_row=ws_sum.max_row, start_column=1, end_row=ws_sum.max_row, end_column=5)
        ws_sum.append([""])
        
        ws_sum.append(["[ ⚠️ 요주의 문항 (평균 6.0점 미만 항목) ]"])
        ws_sum.cell(ws_sum.max_row, 1).font = Font(bold=True, color="FF0000")
        if len(warnings) == 0:
            ws_sum.append(["  • 6.0점 미만인 항목이 없습니다. (모든 문항 우수)"])
            ws_sum.merge_cells(start_row=ws_sum.max_row, start_column=1, end_row=ws_sum.max_row, end_column=5)
        else:
            for idx, val in warnings.items():
                name = idx.split('.', 1)[1] if '.' in idx else idx
                ws_sum.append([f"  • {name} ({val}점) -> 관리자의 원인 분석 및 확인이 필요합니다."])
                ws_sum.merge_cells(start_row=ws_sum.max_row, start_column=1, end_row=ws_sum.max_row, end_column=5)
        ws_sum.append([""])
        
        ws_sum.append(["[ 📌 세부 문항별 평균 만족도 점수 (7점 만점) ]"])
        ws_sum.cell(ws_sum.max_row, 1).font = Font(bold=True)
        headers = ["문항 번호", "평가 항목", "근로자 평균", "실업자 평균", "전체 평균"]
        ws_sum.append(headers)
        table_header_row = ws_sum.max_row
        for col_idx in range(1, 6):
            cell = ws_sum.cell(row=table_header_row, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        has_worker = '근로자 과정' in mean_df.index
        has_unemp = '실업자 과정' in mean_df.index
        for i, col_name in enumerate(numeric_cols, 1):
            w_val = mean_df.loc['근로자 과정', col_name] if has_worker and pd.notna(mean_df.loc['근로자 과정', col_name]) else '-'
            u_val = mean_df.loc['실업자 과정', col_name] if has_unemp and pd.notna(mean_df.loc['실업자 과정', col_name]) else '-'
            o_val = overall_mean[col_name] if pd.notna(overall_mean[col_name]) else '-'
            short_name = col_name.split('.', 1)[1] if '.' in col_name else col_name
            ws_sum.append([i, short_name, w_val, u_val, o_val])
            for col_idx in range(1, 6): ws_sum.cell(row=ws_sum.max_row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

        ws_sum.column_dimensions['A'].width = 10
        ws_sum.column_dimensions['B'].width = 24
        ws_sum.column_dimensions['C'].width = 13
        ws_sum.column_dimensions['D'].width = 13
        ws_sum.column_dimensions['E'].width = 13
        ws_sum.append([""])

        def write_text_section(title, column_name, start_row):
            current_row = start_row
            ws_sum.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=12)
            current_row += 1
            items = final_df[final_df[column_name].astype(str).str.strip() != ''][column_name].dropna().unique()
            if len(items) == 0:
                ws_sum.cell(row=current_row, column=1, value="  • 특별한 의견이 없습니다.")
                ws_sum.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                current_row += 2
            else:
                for item in items:
                    text_str = str(item)
                    cell = ws_sum.cell(row=current_row, column=1, value=f"  • {text_str}")
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                    ws_sum.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                    ws_sum.row_dimensions[current_row].height = max(30, (text_str.count('\n') + (len(text_str) // 35) + 1) * 18)
                    current_row += 1
                current_row += 1
            return current_row
            
        next_row = write_text_section("[ 💡 주요 개선요청사항 ]", "개선요청사항", ws_sum.max_row)
        next_row = write_text_section("[ 💌 주요 수강후기 ]", "수강후기", next_row)

        ws_sum.append(["[ 📋 사후 조치 계획서 (Action Plan) ]"])
        ws_sum.cell(row=ws_sum.max_row, column=1).font = Font(bold=True, size=12)
        ws_sum.append(["구분", "주요 피드백 내용", "개선 및 조치 계획", "", "담당/기한"])
        action_header = ws_sum.max_row
        ws_sum.merge_cells(start_row=action_header, start_column=3, end_row=action_header, end_column=4)
        for col in range(1, 6):
            cell = ws_sum.cell(row=action_header, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in range(1, 6): ws_sum.cell(row=action_header, column=col).border = thin_border

        for i in range(1, 4):
            ws_sum.append([f"{i}", "", "", "", ""])
            curr_row = ws_sum.max_row
            ws_sum.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
            ws_sum.row_dimensions[curr_row].height = 50 
            for col in range(1, 6):
                cell = ws_sum.cell(row=curr_row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws_sum.page_setup.orientation = ws_sum.ORIENTATION_PORTRAIT
        ws_sum.sheet_properties.pageSetUpPr.fitToPage = True
        ws_sum.page_setup.fitToHeight = False
        ws_sum.page_setup.fitToWidth = 1

    return buffer

# --- Streamlit 웹 화면 구성 ---
st.set_page_config(page_title="간호학원 만족도 조사 앱", page_icon="📊", layout="wide")
st.title("📊 간호학원 만족도 조사 대시보드")
st.write("원본 파일(.csv, .xlsx)을 업로드하시면 웹 화면에서 즉시 **요약 인사이트**와 **전체 데이터**를 확인하실 수 있습니다.")

st.divider()
col1, col2 = st.columns(2)
with col1: file_guri = st.file_uploader("📂 구리 학원 결과 업로드", type=['csv', 'xlsx'])
with col2: file_nyj = st.file_uploader("📂 남양주 학원 결과 업로드", type=['csv', 'xlsx'])
st.divider()

if st.button("🚀 데이터 분석 및 엑셀 다운로드 파일 생성", width='stretch'): # ✨ 경고 패치: width 속성 적용
    if file_guri or file_nyj:
        with st.spinner('데이터를 분석하고 웹 대시보드를 생성하고 있습니다...'):
            result_df = process_data(file_guri, file_nyj)
            
            if not result_df.empty:
                # 엑셀 파일은 메모리에 생성해두고
                excel_buffer = generate_excel(result_df)
                
                # 상단에 엑셀 다운로드 버튼 배치
                st.download_button(
                    label="📥 [인쇄용] 완성된 보고서 엑셀 파일 다운로드",
                    data=excel_buffer.getvalue(),
                    file_name="완료_만족도조사_최종보고서(웹연동).xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width='stretch' # ✨ 경고 패치: width 속성 적용
                )
                st.write("") # 간격 띄우기
                
                # --- 웹 대시보드 UI 시작 ---
                tab1, tab2 = st.tabs(["📊 요약 보고서 (인사이트)", "📋 전체 원본 데이터"])
                
                # 숫자 통계 계산 (웹 화면용)
                numeric_cols = [col for col in result_df.columns if col[0].isdigit()]
                df_num = result_df.copy()
                for col in numeric_cols: df_num[col] = pd.to_numeric(df_num[col], errors='coerce')
                
                overall_mean = df_num[numeric_cols].mean().round(2)
                mean_df = df_num.groupby('과정구분')[numeric_cols].mean().round(2)
                
                total_resp = len(result_df)
                overall_avg = overall_mean.mean().round(2)
                recom_cnt = result_df['18.추천여부'].astype(str).str.contains('예').sum()
                recom_rate = round((recom_cnt / total_resp) * 100, 1) if total_resp > 0 else 0

                # [ 탭 1. 요약 보고서 화면 ]
                with tab1:
                    st.subheader("💡 핵심 요약 대시보드")
                    metric1, metric2, metric3 = st.columns(3)
                    metric1.metric("총 응답자 수", f"{total_resp}명")
                    metric2.metric("전체 평균 만족도", f"{overall_avg}점")
                    metric3.metric("지인 추천 의향(율)", f"{recom_rate}%")
                    st.divider()
                    
                    st.subheader("🏆 학원 강점 및 시급한 개선점")
                    col_top, col_bot = st.columns(2)
                    
                    item_means = overall_mean.sort_values(ascending=False)
                    with col_top:
                        st.info("**만족도 Top 3 (강점)**")
                        for i, (idx, val) in enumerate(item_means.head(3).items(), 1):
                            name = idx.split('.', 1)[1] if '.' in idx else idx
                            st.write(f"**{i}위.** {name} ({val}점)")
                            
                    with col_bot:
                        st.error("**만족도 Bottom 3 (약점)**")
                        for i, (idx, val) in enumerate(item_means.tail(3).items(), 1):
                            name = idx.split('.', 1)[1] if '.' in idx else idx
                            st.write(f"**{i}위.** {name} ({val}점)")
                    st.divider()
                    
                    st.subheader("⚠️ 요주의 문항 (평균 6.0점 미만)")
                    warnings = item_means[item_means < 6.0]
                    if len(warnings) == 0:
                        st.success("6.0점 미만인 항목이 없습니다. 전반적으로 훌륭하게 운영되었습니다!")
                    else:
                        for idx, val in warnings.items():
                            name = idx.split('.', 1)[1] if '.' in idx else idx
                            st.warning(f"🚨 **{name} ({val}점)** : 관리자의 원인 분석이 필요합니다.")
                    st.divider()
                    
                    st.subheader("📌 문항별 세부 만족도 점수")
                    table_data = []
                    for i, col_name in enumerate(numeric_cols, 1):
                        w_val = mean_df.loc['근로자 과정', col_name] if '근로자 과정' in mean_df.index and pd.notna(mean_df.loc['근로자 과정', col_name]) else '-'
                        u_val = mean_df.loc['실업자 과정', col_name] if '실업자 과정' in mean_df.index and pd.notna(mean_df.loc['실업자 과정', col_name]) else '-'
                        o_val = overall_mean[col_name] if pd.notna(overall_mean[col_name]) else '-'
                        short_name = col_name.split('.', 1)[1] if '.' in col_name else col_name
                        table_data.append({"문항 번호": i, "평가 항목": short_name, "근로자 평균": w_val, "실업자 평균": u_val, "전체 평균": o_val})
                    
                    # ✨ 혼합 타입 에러 패치: .astype(str) 추가
                    st.dataframe(pd.DataFrame(table_data).astype(str), width='stretch', hide_index=True)
                
                # [ 탭 2. 원본 데이터 화면 ]
                with tab2:
                    st.subheader("📋 전체 원본 데이터 확인")
                    st.write("가로로 스크롤하여 모든 학생들의 개별 응답을 확인할 수 있습니다.")
                    # ✨ 혼합 타입 에러 패치: .astype(str) 추가
                    st.dataframe(result_df.astype(str), width='stretch', hide_index=True)
            else:
                st.error("데이터를 읽어오지 못했습니다. 파일 양식을 확인해 주세요.")
    else:
        st.warning("파일을 업로드해 주세요.")